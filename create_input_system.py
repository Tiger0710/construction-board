"""入力システム生成: ガントチャート（全情報集約）+ 日次入力 + VBA自動同期"""
import os
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

import config

GANTT_DAYS = 31
WEEKDAYS_JP = ["月", "火", "水", "木", "金", "土", "日"]

# ---------- スタイル ----------
NAVY = "1A2E5A"
HEADER_FONT = Font(name="Noto Sans JP", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Noto Sans JP", size=11)
CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
GANTT_DATE_FONT = Font(name="Noto Sans JP", size=9, bold=True)
GANTT_WD_FONT = Font(name="Noto Sans JP", size=9, color="555555")
GANTT_LABEL_FONT = Font(name="Noto Sans JP", size=10, bold=True)
LABEL_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
SAT_FILL = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
SUN_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
BAR_FILL = PatternFill(start_color="93C5FD", end_color="93C5FD", fill_type="solid")

# ガントチャート列定義: (列名, 幅)
GANTT_COLS = [
    ("客先", 12), ("工事件名", 38), ("現場担当者", 10), ("安品担当者", 10),
    ("協力会社名", 16), ("協力会社担当者", 10), ("開始", 10), ("終了", 10),
]
GANTT_LEFT_COLS = len(GANTT_COLS)  # 8列 (A-H)
DATE_START_COL = GANTT_LEFT_COLS + 1  # I列から日付

# ---------- VBA コード ----------
VBA_MODULE = '''\
Public Sub SyncGanttToDaily()
    Dim wsG As Worksheet, wsD As Worksheet

    On Error Resume Next
    Set wsG = ThisWorkbook.Sheets("\u30ac\u30f3\u30c8\u30c1\u30e3\u30fc\u30c8")
    Set wsD = ThisWorkbook.Sheets("\u65e5\u6b21\u5165\u529b")
    On Error GoTo 0

    If wsG Is Nothing Or wsD Is Nothing Then Exit Sub

    Application.ScreenUpdating = False
    Application.EnableEvents = False
    On Error GoTo ErrHandler

    Dim dict As Object
    Set dict = CreateObject("Scripting.Dictionary")

    Dim lastD As Long
    lastD = wsD.Cells(wsD.Rows.Count, 1).End(xlUp).Row
    If lastD < 2 Then lastD = 1

    Dim i As Long
    For i = 2 To lastD
        If Not IsEmpty(wsD.Cells(i, 1).Value) And Not IsEmpty(wsD.Cells(i, 2).Value) Then
            Dim k As String
            k = Format(CDate(wsD.Cells(i, 1).Value), "yyyy-mm-dd") & "|" & _
                Trim(CStr(wsD.Cells(i, 2).Value)) & "|" & Trim(CStr(wsD.Cells(i, 3).Value))
            If Not dict.Exists(k) Then
                dict.Add k, Array( _
                    Trim(CStr(wsD.Cells(i, 4).Value & "")), _
                    Trim(CStr(wsD.Cells(i, 5).Value & "")), _
                    Trim(CStr(wsD.Cells(i, 6).Value & "")))
            End If
        End If
    Next i

    Dim lastG As Long
    lastG = wsG.Cells(wsG.Rows.Count, 1).End(xlUp).Row
    If lastG < 3 Then GoTo Done

    Dim cnt As Long: cnt = 0
    Dim s As Date, e As Date
    For i = 3 To lastG
        If Not IsEmpty(wsG.Cells(i, 1).Value) And Not IsEmpty(wsG.Cells(i, 2).Value) Then
            If IsDate(wsG.Cells(i, 7).Value) And IsDate(wsG.Cells(i, 8).Value) Then
                s = CDate(wsG.Cells(i, 7).Value)
                e = CDate(wsG.Cells(i, 8).Value)
                If e >= s Then cnt = cnt + CLng(e - s) + 1
            End If
        End If
    Next i

    If cnt = 0 Then GoTo Done

    Dim tmpData() As Variant
    ReDim tmpData(1 To cnt, 1 To 6)
    Dim idx As Long: idx = 0

    For i = 3 To lastG
        If Not IsEmpty(wsG.Cells(i, 1).Value) And Not IsEmpty(wsG.Cells(i, 2).Value) Then
            If IsDate(wsG.Cells(i, 7).Value) And IsDate(wsG.Cells(i, 8).Value) Then
                s = CDate(wsG.Cells(i, 7).Value)
                e = CDate(wsG.Cells(i, 8).Value)
                If e >= s Then
                    Dim d As Date
                    Dim cl As String, tt As String
                    cl = Trim(CStr(wsG.Cells(i, 1).Value))
                    tt = Trim(CStr(wsG.Cells(i, 2).Value))
                    For d = s To e
                        idx = idx + 1
                        tmpData(idx, 1) = d
                        tmpData(idx, 2) = cl
                        tmpData(idx, 3) = tt
                        Dim ek As String
                        ek = Format(d, "yyyy-mm-dd") & "|" & cl & "|" & tt
                        If dict.Exists(ek) Then
                            Dim prev As Variant
                            prev = dict(ek)
                            tmpData(idx, 4) = prev(0)
                            tmpData(idx, 5) = prev(1)
                            tmpData(idx, 6) = prev(2)
                        Else
                            tmpData(idx, 4) = ""
                            tmpData(idx, 5) = ""
                            tmpData(idx, 6) = ""
                        End If
                    Next d
                End If
            End If
        End If
    Next i

    Dim clearEnd As Long
    clearEnd = lastD
    If idx + 1 > clearEnd Then clearEnd = idx + 1
    If clearEnd >= 2 Then wsD.Range("A2:F" & clearEnd).ClearContents

    wsD.Range("A2").Resize(idx, 6).Value = tmpData

    Dim rng As Range
    Set rng = wsD.Range("A2:F" & idx + 1)
    rng.Font.Name = "Noto Sans JP"
    rng.Font.Size = 11
    rng.Borders.LineStyle = xlContinuous
    rng.Borders.Weight = xlThin
    rng.VerticalAlignment = xlCenter

    wsD.Range("A2:A" & idx + 1).NumberFormat = "M/D"
    wsD.Range("A2:A" & idx + 1).HorizontalAlignment = xlCenter
    wsD.Range("D2:D" & idx + 1).HorizontalAlignment = xlCenter
    wsD.Range("F2:F" & idx + 1).HorizontalAlignment = xlCenter

    wsD.Range("A1:F" & idx + 1).Sort _
        Key1:=wsD.Range("A2"), Order1:=xlAscending, _
        Key2:=wsD.Range("B2"), Order2:=xlAscending, _
        Key3:=wsD.Range("C2"), Order3:=xlAscending, _
        Header:=xlYes

Done:
    Application.ScreenUpdating = True
    Application.EnableEvents = True
    Exit Sub

ErrHandler:
    Application.ScreenUpdating = True
    Application.EnableEvents = True
End Sub
'''

# ---------- 担当者別サンプル (各20案件) ----------
# (客先, 工事件名, 現場担当者, 安品担当者, 協力会社名, 協力会社担当者, 開始offset, 終了offset)
PROJECTS_BY_PERSON = {
    "担当者A": [
        ("NESIC", "品川基地局アンテナ設置工事", "田中", "佐藤", "共栄電設", "鈴木", 0, 9),
        ("京王", "調布駅構内通信設備更新工事", "田中", "佐藤", "東京通信工業", "伊藤", 0, 3),
        ("小田急", "相模大野駅監視カメラ設置工事", "田中", "中村", "関東電気工事", "小林", 2, 7),
        ("NTT", "港区光ファイバー幹線敷設工事", "田中", "佐藤", "三和通信工業", "森", 1, 12),
        ("KDDI", "世田谷5G基地局新設工事", "田中", "中村", "共栄電設", "高橋", 0, 6),
        ("NESIC", "川崎データセンター電源増設工事", "田中", "佐藤", "大成電気", "吉田", 5, 14),
        ("京王", "府中駅ホームドア通信設備工事", "田中", "中村", "東京通信工業", "渡辺", 3, 8),
        ("ソフトバンク", "新橋局舎無線機器更新工事", "田中", "佐藤", "関東電気工事", "加藤", 7, 15),
        ("小田急", "海老名車両基地信号通信工事", "田中", "中村", "三和通信工業", "佐々木", 4, 10),
        ("NTT", "渋谷マンホール内ケーブル接続工事", "田中", "佐藤", "共栄電設", "山田", 8, 13),
        ("楽天モバイル", "大田区屋上基地局設置工事", "田中", "中村", "大成電気", "松本", 10, 18),
        ("KDDI", "横浜みなとみらいスモールセル設置", "田中", "佐藤", "関東電気工事", "井上", 12, 17),
        ("NESIC", "立川駅北口再開発ビルLAN工事", "田中", "中村", "東京通信工業", "木村", 14, 22),
        ("京王", "聖蹟桜ヶ丘駅放送設備更新", "田中", "佐藤", "三和通信工業", "林", 9, 13),
        ("NTT", "千代田区電話交換機撤去工事", "田中", "中村", "共栄電設", "清水", 16, 21),
        ("ソフトバンク", "品川シーサイドビル構内配線工事", "田中", "佐藤", "大成電気", "山口", 18, 25),
        ("小田急", "本厚木駅Wi-Fi設備設置工事", "田中", "中村", "関東電気工事", "中島", 20, 26),
        ("KDDI", "八王子中継局アンテナ交換工事", "田中", "佐藤", "三和通信工業", "前田", 22, 28),
        ("楽天モバイル", "杉並区電柱共架基地局工事", "田中", "中村", "共栄電設", "小川", 24, 29),
        ("NESIC", "羽田空港第3ターミナル通信設備工事", "田中", "佐藤", "東京通信工業", "岡田", 26, 30),
    ],
    "担当者B": [
        ("NESIC", "横浜駅5G基地局増設工事", "山本", "佐藤", "共栄電設", "高橋", 0, 4),
        ("きんでん", "渋谷再開発ビル電気通信工事", "山本", "佐藤", "大成電気", "吉田", 0, 6),
        ("京王", "新宿駅Wi-Fi設備設置工事", "山本", "中村", "東京通信工業", "渡辺", 3, 8),
        ("NTT", "豊島区光収容局増設工事", "山本", "佐藤", "三和通信工業", "林", 1, 7),
        ("ソフトバンク", "六本木ヒルズ屋上基地局更新", "山本", "中村", "関東電気工事", "加藤", 0, 5),
        ("KDDI", "目黒区マンション共用部通信工事", "山本", "佐藤", "共栄電設", "佐々木", 5, 11),
        ("楽天モバイル", "中野坂上ビル屋上アンテナ設置", "山本", "中村", "大成電気", "山田", 2, 9),
        ("NESIC", "東京駅八重洲地下街Wi-Fi増設", "山本", "佐藤", "東京通信工業", "松本", 7, 14),
        ("きんでん", "新宿三丁目ビル受変電設備更新", "山本", "中村", "三和通信工業", "森", 6, 12),
        ("京王", "橋本駅改札機通信設備工事", "山本", "佐藤", "関東電気工事", "井上", 10, 16),
        ("NTT", "文京区地下管路ケーブル布設工事", "山本", "中村", "共栄電設", "木村", 8, 15),
        ("ソフトバンク", "お台場商業施設DAS設置工事", "山本", "佐藤", "大成電気", "清水", 12, 19),
        ("KDDI", "練馬区鉄塔基地局補強工事", "山本", "中村", "三和通信工業", "山口", 14, 20),
        ("楽天モバイル", "板橋区コンビニ屋上小型基地局", "山本", "佐藤", "関東電気工事", "中島", 16, 22),
        ("NESIC", "品川インターシティ館内放送更新", "山本", "中村", "東京通信工業", "前田", 18, 24),
        ("きんでん", "池袋サンシャイン非常用発電機更新", "山本", "佐藤", "共栄電設", "小川", 13, 19),
        ("京王", "多摩センター駅防犯カメラ増設", "山本", "中村", "大成電気", "岡田", 20, 25),
        ("NTT", "中央区海底ケーブル陸揚局保守", "山本", "佐藤", "三和通信工業", "石井", 22, 27),
        ("ソフトバンク", "二子玉川ライズ通信設備増強", "山本", "中村", "関東電気工事", "斎藤", 24, 29),
        ("KDDI", "調布市鉄塔建替に伴うアンテナ移設", "山本", "佐藤", "共栄電設", "太田", 26, 30),
    ],
    "担当者C": [
        ("小田急", "町田駅構内LAN配線工事", "鈴木", "中村", "関東電気工事", "加藤", 0, 5),
        ("NESIC", "大宮基地局保守点検", "鈴木", "佐藤", "共栄電設", "佐々木", 0, 2),
        ("きんでん", "池袋オフィスビル構内交換設備工事", "鈴木", "中村", "大成電気", "山田", 2, 7),
        ("京王", "八王子駅放送設備更新工事", "鈴木", "中村", "東京通信工業", "松本", 3, 9),
        ("JR東日本", "上野駅構内旅客案内表示器更新", "鈴木", "佐藤", "三和通信工業", "森", 0, 8),
        ("東急", "自由が丘駅ホーム監視カメラ設置", "鈴木", "中村", "関東電気工事", "井上", 1, 6),
        ("NTT", "足立区光クロージャ交換工事", "鈴木", "佐藤", "共栄電設", "木村", 4, 10),
        ("KDDI", "北区王子ビル屋上基地局新設", "鈴木", "中村", "大成電気", "清水", 5, 13),
        ("小田急", "新百合ヶ丘駅構内放送設備更新", "鈴木", "佐藤", "東京通信工業", "山口", 7, 12),
        ("NESIC", "さいたま新都心ビルLAN配線工事", "鈴木", "中村", "三和通信工業", "中島", 6, 14),
        ("JR東日本", "赤羽駅ホームドア通信連動工事", "鈴木", "佐藤", "関東電気工事", "前田", 9, 16),
        ("東急", "武蔵小杉駅改札IC通信設備工事", "鈴木", "中村", "共栄電設", "小川", 10, 15),
        ("きんでん", "有楽町ビル高圧受電設備更新工事", "鈴木", "佐藤", "大成電気", "岡田", 12, 18),
        ("京王", "府中競馬場通信インフラ整備工事", "鈴木", "中村", "東京通信工業", "石井", 14, 20),
        ("NTT", "墨田区マンホール内光接続工事", "鈴木", "佐藤", "三和通信工業", "斎藤", 11, 17),
        ("KDDI", "荒川区商業ビルスモールセル設置", "鈴木", "中村", "関東電気工事", "太田", 16, 22),
        ("小田急", "藤沢駅デジタルサイネージ設置工事", "鈴木", "佐藤", "共栄電設", "藤田", 18, 24),
        ("JR東日本", "浦和駅構内非常通報設備更新", "鈴木", "中村", "大成電気", "三浦", 20, 26),
        ("東急", "中目黒駅構内Wi-Fi設備設置工事", "鈴木", "佐藤", "東京通信工業", "松田", 22, 28),
        ("NESIC", "横須賀基地局鉄塔塗装・点検工事", "鈴木", "中村", "三和通信工業", "金子", 25, 30),
    ],
}

# 日次入力サンプル: (day_offset, proj_idx): (昼/夜, 工事内容, 重点工事)
# ※ 2026-04-01(水)開始 → offset 3=土, 4=日
DAILY_SAMPLES = {
    "担当者A": {
        (0, 0): ("昼", "基礎工事・アンカー打設", "有"),
        (0, 1): ("昼", "機器搬入・仮設置", "無"),
        (0, 3): ("夜", "光ファイバー布設・接続", "有"),
        (0, 4): ("昼", "鉄塔基礎掘削", "無"),
        (1, 0): ("昼", "鉄骨組立・溶接", "有"),
        (1, 1): ("昼", "通信ケーブル敷設", "無"),
        (1, 2): ("昼", "配管ルート確認・墨出し", "無"),
        (1, 3): ("昼", "マンホール内作業・管路清掃", "有"),
        (1, 4): ("昼", "アンテナポール建柱", "無"),
        (2, 0): ("昼", "アンテナブラケット取付", "有"),
        (2, 2): ("昼", "カメラ本体据付・配線", "無"),
        (2, 3): ("昼", "光心線融着接続", "有"),
        (2, 4): ("昼", "基礎コンクリート打設", "有"),
        (3, 0): ("なし", "", ""),
        (3, 1): ("なし", "", ""),
        (3, 3): ("なし", "", ""),
        (3, 4): ("なし", "", ""),
        (4, 0): ("なし", "", ""),
        (4, 3): ("なし", "", ""),
        (4, 4): ("なし", "", ""),
    },
    "担当者B": {
        (0, 0): ("昼", "アンテナ架台組立", "有"),
        (0, 1): ("夜", "ケーブルラック設置", "有"),
        (0, 3): ("昼", "既設機器撤去・搬出", "無"),
        (0, 4): ("昼", "屋上防水養生・墨出し", "無"),
        (1, 0): ("昼", "同軸ケーブル敷設", "無"),
        (1, 1): ("夜", "高圧配線作業", "有"),
        (1, 3): ("昼", "ラック組立・据付", "無"),
        (1, 4): ("昼", "アンテナ基礎アンカー施工", "有"),
        (1, 6): ("昼", "鉄骨補強・溶接作業", "無"),
        (2, 0): ("昼", "コネクタ加工・測定", "無"),
        (2, 1): ("夜", "幹線ケーブル布設", "有"),
        (2, 2): ("昼", "Wi-Fi AP仮設置・電波測定", "無"),
        (2, 6): ("昼", "アンテナ揚重・取付", "有"),
        (3, 0): ("なし", "", ""),
        (3, 1): ("なし", "", ""),
        (3, 3): ("なし", "", ""),
        (3, 4): ("なし", "", ""),
        (4, 0): ("なし", "", ""),
        (4, 1): ("なし", "", ""),
        (4, 4): ("なし", "", ""),
    },
    "担当者C": {
        (0, 0): ("昼", "既設配線撤去", "無"),
        (0, 1): ("昼", "点検・計測作業", "無"),
        (0, 4): ("昼", "案内表示器撤去・仮設", "有"),
        (0, 5): ("昼", "監視カメラブラケット取付", "無"),
        (1, 0): ("昼", "新規配線工事", "無"),
        (1, 1): ("昼", "機器動作確認", "無"),
        (1, 4): ("昼", "表示器取付・配線", "有"),
        (1, 5): ("昼", "カメラ本体据付・調整", "無"),
        (2, 0): ("昼", "HUB設置・パッチケーブル接続", "無"),
        (2, 2): ("夜", "交換機ラック据付", "有"),
        (2, 4): ("昼", "通信試験・調整", "無"),
        (2, 5): ("昼", "映像確認・画角調整", "無"),
        (3, 0): ("なし", "", ""),
        (3, 2): ("なし", "", ""),
        (3, 3): ("なし", "", ""),
        (3, 4): ("なし", "", ""),
        (3, 5): ("なし", "", ""),
        (4, 0): ("なし", "", ""),
        (4, 2): ("なし", "", ""),
        (4, 4): ("なし", "", ""),
        (4, 5): ("なし", "", ""),
    },
}


def apply_header(ws, row, col_start, col_end, fill=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def create_input_file(person_name, projects, daily_data):
    wb = Workbook()
    today = datetime.date.today()

    # ==================== ガントチャート ====================
    ws = wb.active
    ws.title = "ガントチャート"

    # 列幅
    for i, (_, width) in enumerate(GANTT_COLS):
        ws.column_dimensions[get_column_letter(i + 1)].width = width

    # Row 1: ボタン配置エリア + 日付
    ws.row_dimensions[1].height = 28
    for c in range(1, GANTT_LEFT_COLS + 1):
        ws.cell(row=1, column=c).border = THIN_BORDER

    for i in range(GANTT_DAYS):
        d = today + datetime.timedelta(days=i)
        col = DATE_START_COL + i
        c1 = ws.cell(row=1, column=col, value=d)
        c1.number_format = "M/D"
        c1.font = GANTT_DATE_FONT
        c1.alignment = Alignment(horizontal="center")
        c1.border = THIN_BORDER
        if d.weekday() == 5: c1.fill = SAT_FILL
        elif d.weekday() == 6: c1.fill = SUN_FILL
        ws.column_dimensions[get_column_letter(col)].width = 5.5

    # Row 2: ラベル + 曜日
    for i, (label, _) in enumerate(GANTT_COLS):
        cell = ws.cell(row=2, column=i + 1, value=label)
        cell.font = GANTT_LABEL_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        cell.fill = LABEL_FILL

    for i in range(GANTT_DAYS):
        d = today + datetime.timedelta(days=i)
        col = DATE_START_COL + i
        c2 = ws.cell(row=2, column=col, value=WEEKDAYS_JP[d.weekday()])
        c2.font = GANTT_WD_FONT
        c2.alignment = Alignment(horizontal="center")
        c2.border = THIN_BORDER
        if d.weekday() == 5: c2.fill = SAT_FILL
        elif d.weekday() == 6: c2.fill = SUN_FILL

    # データ行 + 空行
    GANTT_NO_WRAP = Alignment(vertical="center", wrap_text=False)
    GANTT_ROW_HEIGHT = 22

    max_data_row = len(projects) + 2 + 5
    for idx, proj in enumerate(projects):
        row = idx + 3
        client, title, our, safety, partner, pp, s_off, e_off = proj
        start_d = today + datetime.timedelta(days=s_off)
        end_d = today + datetime.timedelta(days=e_off)
        vals = [client, title, our, safety, partner, pp, start_d, end_d]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            if c in (3, 4, 6):
                cell.alignment = CENTER_ALIGN
            elif c in (7, 8):
                cell.number_format = "M/D"
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = GANTT_NO_WRAP  # 折り返さない
        for col in range(DATE_START_COL, DATE_START_COL + GANTT_DAYS):
            ws.cell(row=row, column=col).border = THIN_BORDER
        ws.row_dimensions[row].height = GANTT_ROW_HEIGHT

    # 空行
    for r in range(len(projects) + 3, max_data_row + 1):
        for c in range(1, GANTT_LEFT_COLS + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
        for col in range(DATE_START_COL, DATE_START_COL + GANTT_DAYS):
            ws.cell(row=r, column=col).border = THIN_BORDER
        ws.row_dimensions[r].height = GANTT_ROW_HEIGHT

    # 開始/終了の書式
    for r in range(3, max_data_row + 1):
        ws.cell(row=r, column=7).number_format = "M/D"
        ws.cell(row=r, column=8).number_format = "M/D"

    # 条件付き書式: 開始〜終了のバー
    last_col = get_column_letter(DATE_START_COL + GANTT_DAYS - 1)
    fmt_range = f"I3:{last_col}{max_data_row}"
    ws.conditional_formatting.add(fmt_range, FormulaRule(
        formula=['AND(I$1>=$G3,I$1<=$H3)'],
        fill=BAR_FILL,
    ))

    ws.freeze_panes = "I3"

    # ==================== 日次入力 ====================
    ws_d = wb.create_sheet("日次入力")
    d_headers = ["日付", "客先", "工事件名", "昼/夜", "工事内容", "重点工事"]
    d_widths = [12, 12, 28, 8, 40, 10]
    for c, (h, w) in enumerate(zip(d_headers, d_widths), 1):
        ws_d.cell(row=1, column=c, value=h)
        ws_d.column_dimensions[get_column_letter(c)].width = w
    apply_header(ws_d, 1, 1, len(d_headers))

    # ガントから日次入力サンプルを展開
    entries = []
    for idx, proj in enumerate(projects):
        client, title, _, _, _, _, s_off, e_off = proj
        for day in range(s_off, e_off + 1):
            d = today + datetime.timedelta(days=day)
            dn, work, pri = daily_data.get((day, idx), ("", "", ""))
            entries.append((d, client, title, dn, work, pri))
    entries.sort(key=lambda x: (x[0], x[1], x[2]))

    for i, (d, client, title, dn, work, pri) in enumerate(entries, 2):
        ws_d.cell(row=i, column=1, value=d).number_format = "M/D"
        ws_d.cell(row=i, column=2, value=client)
        ws_d.cell(row=i, column=3, value=title)
        ws_d.cell(row=i, column=4, value=dn)
        ws_d.cell(row=i, column=5, value=work)
        ws_d.cell(row=i, column=6, value=pri)
        for c in range(1, 7):
            cell = ws_d.cell(row=i, column=c)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if c in (1, 4, 6) else CELL_ALIGN

    # プルダウン
    dv_dn = DataValidation(type="list", formula1='"昼,夜,なし"', allow_blank=True, showDropDown=False)
    ws_d.add_data_validation(dv_dn)
    dv_dn.add("D2:D500")

    dv_pri = DataValidation(type="list", formula1='"有,無"', allow_blank=True, showDropDown=False)
    ws_d.add_data_validation(dv_pri)
    dv_pri.add("F2:F500")

    ws_d.freeze_panes = "A2"

    # ガントチャートをアクティブに
    wb.active = 0

    xlsx_path = os.path.join(config.DATA_DIR, f"入力_{person_name}.xlsx")
    wb.save(xlsx_path)
    print(f"  作成: {xlsx_path} ({len(projects)}案件)")
    return xlsx_path


def convert_all_to_xlsm(xlsx_paths):
    """全xlsxファイルを一括でxlsm変換 + VBA注入"""
    try:
        import win32com.client
        import pythoncom
    except ImportError:
        print("  win32com未インストール → xlsxのまま (watcher経由で同期)")
        return

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        for xlsx_path in xlsx_paths:
            fname = os.path.basename(xlsx_path)
            xlsm_path = xlsx_path.replace(".xlsx", ".xlsm")
            try:
                wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))

                # 標準モジュール追加 (ボタンから呼ぶ)
                mod = wb.VBProject.VBComponents.Add(1)  # vbext_ct_StdModule
                mod.Name = "GanttSync"
                mod.CodeModule.AddFromString(VBA_MODULE)

                # ガントチャートに「日次入力に反映」ボタン追加 (A1:H1中央)
                ws_gantt = wb.Sheets("\u30ac\u30f3\u30c8\u30c1\u30e3\u30fc\u30c8")
                row1_h = ws_gantt.Rows(1).Height
                header_w = ws_gantt.Cells(1, 8).Left + ws_gantt.Columns(8).Width
                btn_w, btn_h = 170, 22
                btn_left = (header_w - btn_w) / 2
                btn_top = (row1_h - btn_h) / 2
                shp = ws_gantt.Shapes.AddShape(5, btn_left, btn_top, btn_w, btn_h)
                shp.Fill.ForeColor.RGB = 26 + 46 * 256 + 90 * 65536  # Navy
                shp.Line.Visible = False
                tf = shp.TextFrame
                tf.Characters().Text = "\u25b6 \u65e5\u6b21\u5165\u529b\u306b\u53cd\u6620"
                tf.Characters().Font.Color = 16777215  # White
                tf.Characters().Font.Size = 10
                tf.Characters().Font.Bold = True
                tf.HorizontalAlignment = -4108  # xlCenter
                tf.VerticalAlignment = -4108  # xlCenter
                shp.OnAction = "SyncGanttToDaily"

                wb.SaveAs(os.path.abspath(xlsm_path), FileFormat=52)
                wb.Close(False)

                os.remove(xlsx_path)
                print(f"  {fname} → xlsm変換完了")

            except Exception as e:
                err_msg = str(e)
                if "programmatic access" in err_msg.lower() or "信頼性" in err_msg:
                    print(f"  VBA信頼設定が必要 → xlsxのまま")
                    print("  ※ Excel → オプション → トラストセンター → マクロの設定")
                    print("    → 「VBAプロジェクトオブジェクトモデルへのアクセスを信頼する」")
                    break
                else:
                    print(f"  {fname}: xlsm変換失敗 ({e})")

    except Exception as e:
        print(f"  Excel起動エラー: {e}")
    finally:
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
            import time
            time.sleep(1)
        pythoncom.CoUninitialize()


def main():
    os.makedirs(config.DATA_DIR, exist_ok=True)
    print("=== 入力システム生成 ===\n")

    xlsx_paths = []
    for person in PROJECTS_BY_PERSON:
        path = create_input_file(person, PROJECTS_BY_PERSON[person], DAILY_SAMPLES.get(person, {}))
        xlsx_paths.append(path)

    # VBA注入 → xlsm一括変換
    print("\n--- VBAマクロ注入 ---")
    convert_all_to_xlsm(xlsx_paths)

    print(f"\n完了! {config.DATA_DIR}")
    print("\n動作: ガントチャート編集 → 保存 → 日次入力自動更新 → HTML自動生成")


if __name__ == "__main__":
    main()
