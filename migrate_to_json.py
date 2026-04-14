"""既存Excel入力ファイルを新JSON形式に変換するワンショットスクリプト

Usage:
    python migrate_to_json.py [--month 2604] [--input-dir "E:\★工事共有\工事予定表"]
"""
import os
import sys
import json
import hashlib
import datetime
import argparse
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def _to_date(val):
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, str):
        try:
            return datetime.date.fromisoformat(val.replace("/", "-").strip())
        except ValueError:
            return None
    return None


def gen_id(client, title):
    """client+titleから安定したIDを生成"""
    raw = f"{client}|{title}"
    return hashlib.md5(raw.encode()).hexdigest()[:8]


def read_input_file(fpath):
    """1ファイルからproject情報とdailyデータを抽出"""
    fname = os.path.basename(fpath)
    print(f"  読み込み: {fname}")
    wb = load_workbook(fpath, read_only=True, data_only=True)

    projects = []
    project_ids = {}

    # ガントチャート読み込み
    if "ガントチャート" in wb.sheetnames:
        for row in wb["ガントチャート"].iter_rows(min_row=3, values_only=True):
            if not row or not row[0] or not row[1]:
                continue
            client = str(row[0]).strip()
            title = str(row[1]).strip()
            pid = gen_id(client, title)
            project_ids[(client, title)] = pid

            start = _to_date(row[6]) if len(row) > 6 else None
            end = _to_date(row[7]) if len(row) > 7 else None
            if not start or not end:
                continue

            projects.append({
                "id": pid,
                "client": client,
                "title": title,
                "our_person": str(row[2] or "").strip() if len(row) > 2 else "",
                "safety_person": str(row[3] or "").strip() if len(row) > 3 else "",
                "partner": str(row[4] or "").strip() if len(row) > 4 else "",
                "partner_person": str(row[5] or "").strip() if len(row) > 5 else "",
                "start_date": start.isoformat(),
                "end_date": end.isoformat(),
            })

    # 日次入力読み込み
    daily = {}
    if "日次入力" in wb.sheetnames:
        ws = wb["日次入力"]
        header_d = str(ws.cell(row=1, column=4).value or "").strip()
        is_split = header_d == "昼"
        is_new = header_d == "現場担当者"

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[1] or not row[2]:
                continue
            date_val = _to_date(row[0])
            if not date_val:
                continue
            client = str(row[1]).strip()
            title = str(row[2]).strip()
            pid = project_ids.get((client, title))
            if not pid:
                pid = gen_id(client, title)
                project_ids[(client, title)] = pid

            key = f"{pid}/{date_val.isoformat()}"

            if is_split:
                # 15列昼夜分離形式
                day_val = str(row[3] or "").strip() if len(row) > 3 else "昼"
                night_val = str(row[9] or "").strip() if len(row) > 9 else ""
                daily[key] = {
                    "day": day_val != "なし",
                    "day_work": str(row[6] or "").strip() if len(row) > 6 else "",
                    "day_priority": str(row[7] or "").strip() if len(row) > 7 else "",
                    "day_priority_detail": str(row[8] or "").strip() if len(row) > 8 else "",
                    "day_our_person": str(row[4] or "").strip() if len(row) > 4 else "",
                    "day_safety_person": str(row[5] or "").strip() if len(row) > 5 else "",
                    "night": night_val == "夜",
                    "night_work": str(row[12] or "").strip() if len(row) > 12 else "",
                    "night_priority": str(row[13] or "").strip() if len(row) > 13 else "",
                    "night_priority_detail": str(row[14] or "").strip() if len(row) > 14 else "",
                    "night_our_person": str(row[10] or "").strip() if len(row) > 10 else "",
                    "night_safety_person": str(row[11] or "").strip() if len(row) > 11 else "",
                }
            elif is_new:
                # 9列新形式
                day_night = str(row[5] or "").strip() if len(row) > 5 else ""
                is_night = day_night == "夜"
                daily[key] = {
                    "day": not is_night and day_night != "なし",
                    "day_work": str(row[6] or "").strip() if len(row) > 6 and not is_night else "",
                    "day_priority": str(row[7] or "").strip() if len(row) > 7 and not is_night else "",
                    "day_priority_detail": str(row[8] or "").strip() if len(row) > 8 and not is_night else "",
                    "day_our_person": str(row[3] or "").strip() if len(row) > 3 and not is_night else "",
                    "day_safety_person": str(row[4] or "").strip() if len(row) > 4 and not is_night else "",
                    "night": is_night,
                    "night_work": str(row[6] or "").strip() if len(row) > 6 and is_night else "",
                    "night_priority": str(row[7] or "").strip() if len(row) > 7 and is_night else "",
                    "night_priority_detail": str(row[8] or "").strip() if len(row) > 8 and is_night else "",
                    "night_our_person": str(row[3] or "").strip() if len(row) > 3 and is_night else "",
                    "night_safety_person": str(row[4] or "").strip() if len(row) > 4 and is_night else "",
                }
            else:
                # 6列旧形式
                day_night = str(row[3] or "").strip() if len(row) > 3 else ""
                is_night = day_night == "夜"
                daily[key] = {
                    "day": not is_night and day_night != "なし",
                    "day_work": str(row[4] or "").strip() if len(row) > 4 and not is_night else "",
                    "day_priority": str(row[5] or "").strip() if len(row) > 5 and not is_night else "",
                    "day_priority_detail": "",
                    "day_our_person": "",
                    "day_safety_person": "",
                    "night": is_night,
                    "night_work": str(row[4] or "").strip() if len(row) > 4 and is_night else "",
                    "night_priority": str(row[5] or "").strip() if len(row) > 5 and is_night else "",
                    "night_priority_detail": "",
                    "night_our_person": "",
                    "night_safety_person": "",
                }

    wb.close()

    print(f"    案件: {len(projects)}件, 日次: {len(daily)}件")
    return projects, daily


def clean_daily(daily):
    """デフォルト値のみのdailyエントリを除去"""
    cleaned = {}
    for key, dd in daily.items():
        is_default = (
            dd.get("day", True) is True
            and not dd.get("night", False)
            and not dd.get("day_work", "")
            and not dd.get("day_priority", "")
            and not dd.get("day_priority_detail", "")
            and not dd.get("day_our_person", "")
            and not dd.get("day_safety_person", "")
            and not dd.get("night_work", "")
            and not dd.get("night_priority", "")
            and not dd.get("night_priority_detail", "")
            and not dd.get("night_our_person", "")
            and not dd.get("night_safety_person", "")
        )
        if not is_default:
            cleaned[key] = dd
    return cleaned


def main():
    parser = argparse.ArgumentParser(description="Excel → JSON移行")
    parser.add_argument("--month", default="2604", help="対象月 (YYMM)")
    parser.add_argument("--input-dir", default=r"E:\★工事共有\工事予定表",
                        help="入力ディレクトリ")
    parser.add_argument("--output", default=None, help="出力先 (デフォルト: data/{month}.json)")
    args = parser.parse_args()

    month_dir = os.path.join(args.input_dir, args.month)
    if not os.path.isdir(month_dir):
        print(f"エラー: ディレクトリが見つかりません: {month_dir}")
        return 1

    print(f"=== Excel → JSON 移行 ({args.month}) ===\n")
    print(f"入力: {month_dir}")

    # 入力ファイル検索
    import glob
    files = glob.glob(os.path.join(month_dir, "入力_*.xlsm"))
    files += glob.glob(os.path.join(month_dir, "入力_*.xlsx"))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    # 「編集中」を除外
    files = [f for f in files if "編集中" not in os.path.basename(f)]

    if not files:
        print("入力ファイルなし")
        return 1

    print(f"ファイル数: {len(files)}\n")

    all_projects = {}  # id → project (重複排除)
    all_daily = {}

    for fpath in sorted(files):
        projects, daily = read_input_file(fpath)
        for proj in projects:
            all_projects[proj["id"]] = proj
        all_daily.update(daily)

    # デフォルト値のdailyを除去
    all_daily = clean_daily(all_daily)

    result = {
        "projects": list(all_projects.values()),
        "daily": all_daily,
    }

    output = args.output or os.path.join(BASE_DIR, f"_{args.month}.json")
    os.makedirs(os.path.dirname(output), exist_ok=True)
    with open(output, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\n出力: {output}")
    print(f"  案件: {len(result['projects'])}件")
    print(f"  日次: {len(result['daily'])}件")
    print("\n完了!")
    return 0


if __name__ == "__main__":
    sys.exit(main())
