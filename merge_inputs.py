"""入力ファイル統合: ガントチャートから自動生成 + 日次入力オーバーレイ → 工事予定表.xlsx

ファイル名パターン:
  入力_{担当者名}_{YYMM}.xlsm  → 月別ファイル (2604 = 2026年4月)
  入力_{担当者名}.xlsm          → 月指定なし (常に読み込み)

ガントチャートの開始〜終了日から自動で日次エントリーを生成。
日次入力シートに手入力データがあればそちらを優先。
→ VBAマクロ不要。ガントチャートを埋めるだけで動く。
"""
import os
import re
import sys
import glob
import datetime
from collections import Counter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config

# ファイル名から月(YYMM)を抽出するパターン
MONTH_RE = re.compile(r"^入力_.+_(\d{4})\.(xlsx|xlsm)$")


# ---------- スタイル ----------
NAVY = "1A2E5A"
HEADER_FONT = Font(name="Noto Sans JP", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Noto Sans JP", size=11)
CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
PRIORITY_FILL = PatternFill(start_color="F0F3FA", end_color="F0F3FA", fill_type="solid")
NIGHT_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")


def _to_date(val):
    """各種日付型 → datetime.date に変換"""
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, (int, float)):
        return None
    if isinstance(val, str):
        try:
            return datetime.date.fromisoformat(val.replace("/", "-").strip())
        except ValueError:
            return None
    return None


def load_input_file(fpath):
    """1ファイルからガントチャート + 日次入力を読み込み、エントリーを自動生成

    ガントチャートの開始〜終了日から全日分のエントリーを生成し、
    日次入力シートの手入力データがあればオーバーレイする。
    """
    fname = os.path.basename(fpath)
    wb = load_workbook(fpath, read_only=True, data_only=True)

    # ガントチャート: 案件情報 + 日付範囲
    # A=客先, B=工事件名, C=現場担当者, D=安品担当者, E=協力会社名, F=協力会社担当者, G=開始, H=終了
    project_info = {}
    gantt_ranges = []

    if "ガントチャート" in wb.sheetnames:
        for row in wb["ガントチャート"].iter_rows(min_row=3, values_only=True):
            if not row or not row[0] or not row[1]:
                continue
            client = str(row[0]).strip()
            title = str(row[1]).strip()
            key = (client, title)
            project_info[key] = {
                "our_person": str(row[2] or "").strip() if len(row) > 2 else "",
                "safety_person": str(row[3] or "").strip() if len(row) > 3 else "",
                "partner": str(row[4] or "").strip() if len(row) > 4 else "",
                "partner_person": str(row[5] or "").strip() if len(row) > 5 else "",
            }
            start = _to_date(row[6]) if len(row) > 6 else None
            end = _to_date(row[7]) if len(row) > 7 else None
            if start and end and end >= start:
                gantt_ranges.append((client, title, start, end))

    # 日次入力: 手入力データ (lookup用)
    # A=日付, B=客先, C=工事件名, D=昼/夜, E=工事内容, F=重点工事
    daily_dict = {}
    if "日次入力" in wb.sheetnames:
        for row in wb["日次入力"].iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[1] or not row[2]:
                continue
            date_val = _to_date(row[0])
            if not date_val:
                continue
            client = str(row[1]).strip()
            title = str(row[2]).strip()
            day_night = str(row[3] or "").strip() if len(row) > 3 else ""
            work_content = str(row[4] or "").strip() if len(row) > 4 else ""
            priority = str(row[5] or "").strip() if len(row) > 5 else ""
            daily_dict[(date_val, client, title)] = {
                "day_night": day_night,
                "work_content": work_content,
                "priority": priority,
            }

    wb.close()

    # ガントの日付範囲からエントリー自動生成 + 日次入力オーバーレイ
    entries = []
    for client, title, start, end in gantt_ranges:
        d = start
        while d <= end:
            user = daily_dict.pop((d, client, title), None)
            if user:
                entries.append({
                    "date": d, "client": client, "title": title,
                    "day_night": user["day_night"] or "昼",
                    "work_content": user["work_content"],
                    "priority": user["priority"],
                    "source_file": fname,
                })
            else:
                entries.append({
                    "date": d, "client": client, "title": title,
                    "day_night": "昼",
                    "work_content": "",
                    "priority": "",
                    "source_file": fname,
                })
            d += datetime.timedelta(days=1)

    # ガントに紐付かない日次入力（残り）も拾う
    for (d, client, title), user in daily_dict.items():
        entries.append({
            "date": d, "client": client, "title": title,
            "day_night": user["day_night"],
            "work_content": user["work_content"],
            "priority": user["priority"],
            "source_file": fname,
        })

    return project_info, entries


def get_target_months():
    """読み込み対象の月(YYMM)リストを返す (当月 + 翌月)"""
    today = datetime.date.today()
    current = f"{today.year % 100:02d}{today.month:02d}"
    next_month = today.replace(day=28) + datetime.timedelta(days=4)
    next_m = f"{next_month.year % 100:02d}{next_month.month:02d}"
    return [current, next_m]


def filter_by_month(files):
    """月別ファイルを当月+翌月でフィルタ。月指定なしファイルは常に含む"""
    target_months = get_target_months()
    result = []
    for f in files:
        fname = os.path.basename(f)
        m = MONTH_RE.match(fname)
        if m:
            if m.group(1) in target_months:
                result.append(f)
        else:
            result.append(f)
    return result


def merge_all():
    """全ファイル読み込み → ガントチャートJOIN → 統合リスト"""
    files = []
    for d in (config.INPUT_DIR, config.DATA_DIR):
        if not os.path.isdir(d):
            continue
        for ext in ("xlsx", "xlsm"):
            files.extend(glob.glob(os.path.join(d, f"入力_*.{ext}")))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]

    files = filter_by_month(files)

    # 重複除去 (同名ファイルはINPUT_DIR優先)
    seen = {}
    for f in files:
        name = os.path.basename(f)
        if name not in seen or f.startswith(config.INPUT_DIR):
            seen[name] = f
    files = list(seen.values())

    if not files:
        print("警告: 入力ファイルが見つかりません")
        return []

    all_merged = []

    for fpath in sorted(files):
        fname = os.path.basename(fpath)
        print(f"  読み込み: {fname}")
        try:
            project_info, entries = load_input_file(fpath)
            gantt_count = len(project_info)
            print(f"    案件: {gantt_count}件, エントリー: {len(entries)}件")

            for e in entries:
                if e["day_night"] == "なし":
                    continue
                key = (e["client"], e["title"])
                info = project_info.get(key, {})
                all_merged.append({
                    "date": e["date"],
                    "client": e["client"],
                    "title": e["title"],
                    "our_person": info.get("our_person", ""),
                    "safety_person": info.get("safety_person", ""),
                    "partner": info.get("partner", ""),
                    "partner_person": info.get("partner_person", ""),
                    "work_content": e["work_content"],
                    "work_time": e["day_night"],
                    "priority": e["priority"],
                })
        except Exception as e:
            print(f"    エラー: {e}")

    all_merged.sort(key=lambda x: (
        x["date"],
        0 if x["priority"] == "有" else 1,
        x["title"],
    ))
    return all_merged


def write_output(merged, dry_run=False):
    """統合結果 → 工事予定表.xlsx"""
    if dry_run:
        print(f"\n[dry-run] {len(merged)}件:")
        for item in merged:
            night = " [夜間]" if item["work_time"] == "夜" else ""
            pri = " ★" if item["priority"] == "有" else ""
            work = item["work_content"][:15] or "(未記入)"
            print(f"  {item['date']} | {item['client']} | {item['title'][:18]} | {item['work_time']}{night} | {work}{pri}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "工事予定"

    headers = ["日付", "客先", "工事件名", "弊社担当者", "安品担当者",
               "協力会社名", "協力会社担当者", "作業内容", "昼/夜", "重点作業"]
    widths = [14, 12, 30, 10, 10, 18, 10, 40, 8, 10]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(c)].width = w

    for i, item in enumerate(merged, 2):
        vals = [item["date"], item["client"], item["title"],
                item["our_person"], item["safety_person"],
                item["partner"], item["partner_person"],
                item["work_content"], item["work_time"], item["priority"]]
        is_night = item["work_time"] == "夜"
        is_pri = item["priority"] == "有"

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            if c == 1:
                cell.number_format = "YYYY/MM/DD"
                cell.alignment = CENTER_ALIGN
            elif c in (4, 5, 7, 9, 10):
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = CELL_ALIGN
            if is_night:
                cell.fill = NIGHT_FILL
            elif is_pri:
                cell.fill = PRIORITY_FILL

    ws.freeze_panes = "A2"
    wb.save(config.EXCEL_PATH)

    print(f"\n保存: {config.EXCEL_PATH}")
    print(f"  合計: {len(merged)}件")
    counts = Counter(
        item["date"].isoformat() if hasattr(item["date"], "isoformat") else str(item["date"])
        for item in merged
    )
    for d in sorted(counts):
        print(f"  {d}: {counts[d]}件")


def main():
    print("=== 工事予定表 統合 ===\n")
    dry_run = "--dry-run" in sys.argv

    print("[1/2] 入力ファイル読み込み...")
    merged = merge_all()
    print(f"\n  統合: {len(merged)}件")

    if not merged:
        print("\nデータなし。終了。")
        return

    print(f"\n[2/2] {'プレビュー' if dry_run else '書き出し'}...")
    write_output(merged, dry_run=dry_run)

    if not dry_run:
        print("\n完了!")


if __name__ == "__main__":
    main()
