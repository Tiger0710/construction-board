"""ガントチャート(開始/終了) → 日次入力シート自動生成"""
import os
import glob
import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

import config

NAVY = "1A2E5A"
HEADER_FONT = Font(name="Noto Sans JP", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Noto Sans JP", size=11)
CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def to_date(val):
    """セル値をdateに変換"""
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, str):
        try:
            return datetime.date.fromisoformat(val.replace("/", "-").strip())
        except ValueError:
            return None
    return None


def read_gantt(ws):
    """ガントチャート読み込み
    A=客先, B=工事件名, G=開始, H=終了
    → 開始〜終了の各日に (date, 客先, 工事件名) を展開
    """
    entries = []
    for row in range(3, ws.max_row + 1):
        client = ws.cell(row=row, column=1).value
        title = ws.cell(row=row, column=2).value
        if not client or not title:
            continue
        client = str(client).strip()
        title = str(title).strip()

        start = to_date(ws.cell(row=row, column=7).value)
        end = to_date(ws.cell(row=row, column=8).value)
        if not start or not end or end < start:
            continue

        d = start
        while d <= end:
            entries.append((d, client, title))
            d += datetime.timedelta(days=1)

    entries.sort(key=lambda x: (x[0], x[1], x[2]))
    return entries


def read_existing_daily(ws):
    """既存の日次入力を読み込み
    A=日付, B=客先, C=工事件名, D=昼/夜, E=工事内容, F=重点工事
    → {(date_iso, 客先, 工事件名): {day_night, work_content, priority}}
    """
    existing = {}
    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row, column=1).value
        client = ws.cell(row=row, column=2).value
        title = ws.cell(row=row, column=3).value
        if not date_val or not client or not title:
            continue
        d = to_date(date_val)
        if not d:
            continue
        key = (d.isoformat(), str(client).strip(), str(title).strip())
        existing[key] = {
            "day_night": str(ws.cell(row=row, column=4).value or "").strip(),
            "work_content": str(ws.cell(row=row, column=5).value or "").strip(),
            "priority": str(ws.cell(row=row, column=6).value or "").strip(),
        }
    return existing


def sync_file(fpath):
    """1ファイルのガント → 日次入力を同期"""
    fname = os.path.basename(fpath)
    wb = load_workbook(fpath)

    if "ガントチャート" not in wb.sheetnames:
        print(f"  スキップ: {fname} (ガントチャートなし)")
        wb.close()
        return

    # ガント → 日次エントリ展開
    gantt_entries = read_gantt(wb["ガントチャート"])

    # 既存の手入力データを保持
    existing = {}
    if "日次入力" in wb.sheetnames:
        existing = read_existing_daily(wb["日次入力"])

    # 日次入力シートがなければ作成
    if "日次入力" not in wb.sheetnames:
        ws_d = wb.create_sheet("日次入力")
        headers = ["日付", "客先", "工事件名", "昼/夜", "工事内容", "重点工事"]
        widths = [12, 12, 28, 8, 40, 10]
        for c, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws_d.cell(row=1, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            ws_d.column_dimensions[get_column_letter(c)].width = w

        dv_dn = DataValidation(type="list", formula1='"昼,夜,なし"', allow_blank=True, showDropDown=False)
        ws_d.add_data_validation(dv_dn)
        dv_dn.add("D2:D500")

        dv_pri = DataValidation(type="list", formula1='"有,無"', allow_blank=True, showDropDown=False)
        ws_d.add_data_validation(dv_pri)
        dv_pri.add("F2:F500")

        ws_d.freeze_panes = "A2"
    else:
        ws_d = wb["日次入力"]

    # Row 2以降クリア
    for row in range(2, max(ws_d.max_row + 1, 2)):
        for col in range(1, 7):
            ws_d.cell(row=row, column=col).value = None

    # 書き込み
    preserved = 0
    for i, (d, client, title) in enumerate(gantt_entries, 2):
        key = (d.isoformat(), client, title)
        prev = existing.get(key, {})
        if prev.get("work_content") or prev.get("day_night"):
            preserved += 1

        ws_d.cell(row=i, column=1, value=d).number_format = "M/D"
        ws_d.cell(row=i, column=2, value=client)
        ws_d.cell(row=i, column=3, value=title)
        ws_d.cell(row=i, column=4, value=prev.get("day_night", ""))
        ws_d.cell(row=i, column=5, value=prev.get("work_content", ""))
        ws_d.cell(row=i, column=6, value=prev.get("priority", ""))
        for c in range(1, 7):
            cell = ws_d.cell(row=i, column=c)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if c in (1, 4, 6) else CELL_ALIGN

    wb.save(fpath)
    print(f"  同期: {fname} → {len(gantt_entries)}件 (既存保持: {preserved}件)")


def main():
    print("=== ガントチャート → 日次入力 同期 ===\n")
    files = [f for f in glob.glob(os.path.join(config.DATA_DIR, "入力_*.xlsx"))
             if not os.path.basename(f).startswith("~$")]

    if not files:
        print("入力ファイルが見つかりません")
        return

    for fpath in sorted(files):
        try:
            sync_file(fpath)
        except Exception as e:
            print(f"  エラー: {os.path.basename(fpath)} - {e}")

    print("\n完了! 日次入力の昼/夜・工事内容・重点工事を記入してください。")


if __name__ == "__main__":
    main()
