"""工事予定表Excelテンプレート生成スクリプト"""
import os
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

import config


def create_template():
    """工事予定表テンプレートExcelを生成する"""
    wb = Workbook()
    ws = wb.active
    ws.title = "工事予定"

    # ── ヘッダー ──
    headers = ["日付", "開始時刻", "終了時刻", "工事名", "場所", "担当者", "ステータス", "進捗(%)", "備考"]
    header_fill = PatternFill(start_color="1A2040", end_color="1A2040", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # ── 列幅 ──
    col_widths = {"A": 12, "B": 10, "C": 10, "D": 30, "E": 20, "F": 12, "G": 12, "H": 10, "I": 30}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── DataValidation: ステータス(G列) ──
    status_dv = DataValidation(
        type="list",
        formula1='"予定,準備中,進行中,完了,遅延,中止"',
        allow_blank=True,
    )
    status_dv.error = "予定, 準備中, 進行中, 完了, 遅延, 中止 から選択してください"
    status_dv.errorTitle = "無効なステータス"
    status_dv.prompt = "ステータスを選択"
    status_dv.promptTitle = "ステータス"
    ws.add_data_validation(status_dv)
    status_dv.add("G2:G1048576")

    # ── DataValidation: 進捗(H列) ──
    progress_dv = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="100",
        allow_blank=True,
    )
    progress_dv.error = "0〜100の整数を入力してください"
    progress_dv.errorTitle = "無効な進捗値"
    progress_dv.prompt = "0〜100の整数"
    progress_dv.promptTitle = "進捗(%)"
    ws.add_data_validation(progress_dv)
    progress_dv.add("H2:H1048576")

    # ── サンプルデータ (50件) ──
    D = datetime.date(2026, 3, 18)
    T = datetime.time
    sample_data = [
        # === 早朝帯 06:00-09:00（完了が多い） ===
        (D, T(6, 0), T(7, 30), "受水槽清掃", "B1機械室", "田中", "完了", 100, "年次点検"),
        (D, T(6, 0), T(8, 0), "高圧受電設備点検", "屋上キュービクル", "佐藤", "完了", 100, "停電作業 06:00-08:00"),
        (D, T(6, 30), T(7, 30), "駐車場ライン引き直し", "駐車場A", "鈴木", "完了", 100, "早朝施工"),
        (D, T(6, 30), T(9, 0), "給排水管洗浄", "本館B1配管室", "高橋", "完了", 100, "断水あり 06:30-09:00"),
        (D, T(7, 0), T(8, 0), "消防設備点検", "本館1F", "渡辺", "完了", 100, "消火器・感知器確認"),
        (D, T(7, 0), T(9, 0), "非常階段手摺り修繕", "東棟非常階段", "伊藤", "完了", 100, "塗装完了"),
        (D, T(7, 30), T(9, 30), "照明LED化工事", "本館2F廊下", "山本", "完了", 100, "蛍光灯40本交換"),
        (D, T(8, 0), T(9, 0), "自動ドア定期点検", "本館1Fエントランス", "中村", "完了", 100, "センサー調整済み"),
        (D, T(8, 0), T(10, 0), "外壁タイル補修", "東棟外壁", "小林", "完了", 100, "足場設置済み"),
        (D, T(8, 30), T(9, 30), "館内放送設備点検", "本館1F管理室", "加藤", "完了", 100, "定期点検"),
        # === 午前帯 09:00-12:00（進行中が混ざる） ===
        (D, T(9, 0), T(11, 0), "1Fエントランス改修", "本館1F", "田中", "進行中", 60, "塗装作業中"),
        (D, T(9, 0), T(12, 0), "空調配管交換", "3F機械室", "佐藤", "進行中", 45, "冷媒回収完了・配管切断中"),
        (D, T(9, 0), T(10, 30), "電気設備点検", "別館2F電気室", "鈴木", "進行中", 70, "分電盤確認中"),
        (D, T(9, 30), T(11, 30), "トイレ改修工事", "本館3Fトイレ", "高橋", "進行中", 55, "使用不可・仮設トイレ設置"),
        (D, T(9, 30), T(12, 30), "屋上防水シート補修", "本館屋上", "渡辺", "進行中", 30, "雨天延期あり"),
        (D, T(10, 0), T(11, 0), "火災報知器交換", "別館1F", "伊藤", "進行中", 80, "残り3台"),
        (D, T(10, 0), T(12, 0), "ガス配管漏れ修繕", "本館B1ボイラー室", "山本", "遅延", 40, "部品入荷待ち"),
        (D, T(10, 30), T(12, 30), "窓ガラスフィルム貼付", "本館5F", "中村", "進行中", 50, "遮熱フィルム"),
        (D, T(11, 0), T(12, 0), "排煙設備点検", "東棟3F", "小林", "完了", 100, "問題なし"),
        (D, T(11, 0), T(13, 0), "エレベーター定期点検", "エレベーターA号機", "加藤", "進行中", 65, "B号機は午後実施"),
        # === 午後帯 12:00-18:00（予定・準備中が多い） ===
        (D, T(12, 30), T(14, 0), "カーペット張替え", "3F会議室", "田中", "準備中", 10, "家具移動中"),
        (D, T(13, 0), T(15, 0), "空調設備点検", "別館3F", "佐藤", "予定", 0, "定期点検"),
        (D, T(13, 0), T(14, 30), "スプリンクラー点検", "本館全フロア", "鈴木", "予定", 0, "放水テストあり"),
        (D, T(13, 0), T(16, 0), "外壁塗装", "西棟外壁", "高橋", "準備中", 5, "足場組立中"),
        (D, T(13, 30), T(15, 30), "分電盤交換", "別館B1", "渡辺", "予定", 0, "停電あり 13:30-15:30"),
        (D, T(14, 0), T(15, 0), "避難誘導灯交換", "本館全階段", "伊藤", "予定", 0, "LED化"),
        (D, T(14, 0), T(16, 0), "駐車場アスファルト補修", "駐車場B", "山本", "予定", 0, "通行止め"),
        (D, T(14, 0), T(17, 0), "防水工事", "屋上", "中村", "準備中", 10, "雨天時は延期"),
        (D, T(14, 30), T(16, 30), "OAフロア配線整理", "本館4Fオフィス", "小林", "予定", 0, "テナント休業中に実施"),
        (D, T(15, 0), T(16, 0), "貯水槽水質検査", "本館屋上", "加藤", "予定", 0, "法定検査"),
        (D, T(15, 0), T(17, 0), "内装クロス張替え", "別館2F応接室", "田中", "予定", 0, "騒音注意"),
        (D, T(15, 30), T(17, 0), "ボイラー点検", "B1ボイラー室", "佐藤", "予定", 0, "給湯停止あり"),
        (D, T(16, 0), T(17, 30), "防犯カメラ増設", "駐車場A出入口", "鈴木", "予定", 0, "配線工事"),
        (D, T(16, 0), T(18, 0), "共用部照明交換", "本館1F-5F廊下", "高橋", "中止", 0, "資材未着のため中止"),
        (D, T(16, 30), T(18, 0), "自家発電設備点検", "屋上発電機室", "渡辺", "予定", 0, "年次負荷試験"),
        (D, T(17, 0), T(18, 0), "排水ポンプ交換", "B1排水ピット", "伊藤", "予定", 0, "排水制限あり"),
        (D, T(17, 0), T(18, 30), "段差解消スロープ設置", "本館1F裏口", "山本", "遅延", 25, "設計変更対応中"),
        (D, T(17, 0), T(19, 0), "屋上フェンス取替", "別館屋上", "中村", "予定", 0, "安全帯着用"),
        (D, T(17, 30), T(18, 30), "エレベーター定期点検", "エレベーターB号機", "加藤", "準備中", 5, "A号機完了後に着手"),
        (D, T(17, 30), T(19, 30), "漏水調査", "本館3F天井裏", "小林", "予定", 0, "赤外線カメラ使用"),
        # === 夕方以降 18:00-21:00（ほぼ予定） ===
        (D, T(18, 0), T(19, 0), "機械式駐車場点検", "駐車場C地下", "田中", "予定", 0, "利用停止 18:00-19:00"),
        (D, T(18, 0), T(20, 0), "グリストラップ清掃", "本館B1厨房下", "佐藤", "予定", 0, "閉店後作業"),
        (D, T(18, 30), T(20, 30), "床ワックス塗布", "本館1Fロビー", "鈴木", "予定", 0, "乾燥まで通行止め"),
        (D, T(19, 0), T(20, 0), "非常用照明バッテリー交換", "別館全フロア", "高橋", "予定", 0, "定期交換"),
        (D, T(19, 0), T(21, 0), "空調ダクト清掃", "本館2F-4F", "渡辺", "予定", 0, "騒音注意・営業時間外"),
        (D, T(19, 30), T(21, 00), "配管保温材巻き直し", "東棟B1配管室", "伊藤", "予定", 0, "アスベスト含有なし確認済"),
        (D, T(20, 0), T(21, 0), "セキュリティゲート調整", "本館1Fエントランス", "山本", "予定", 0, "カード認証テスト"),
        (D, T(20, 0), T(21, 0), "汚水管高圧洗浄", "別館B1", "中村", "遅延", 20, "前工程遅延の影響"),
        (D, T(20, 30), T(21, 0), "蓄電池設備点検", "屋上蓄電池室", "小林", "予定", 0, "BCP関連設備"),
        (D, T(20, 30), T(21, 0), "守衛室モニター交換", "本館1F守衛室", "加藤", "予定", 0, "4Kモニター2台"),
    ]

    data_font = Font(size=11)
    date_alignment = Alignment(horizontal="center", vertical="center")
    num_alignment = Alignment(horizontal="right", vertical="center")

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

            # 日付・時刻列は中央揃え
            if col_idx <= 3:
                cell.alignment = date_alignment
                if col_idx == 1:
                    cell.number_format = "YYYY/MM/DD"
                else:
                    cell.number_format = "HH:MM"
            # 進捗列は右揃え
            elif col_idx == 8:
                cell.alignment = num_alignment
            else:
                cell.alignment = Alignment(vertical="center")

    # ── 条件付き書式: ステータス列(G列) ──
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    ws.conditional_formatting.add(
        "G2:G1048576",
        CellIsRule(operator="equal", formula=['"進行中"'], fill=yellow_fill),
    )
    ws.conditional_formatting.add(
        "G2:G1048576",
        CellIsRule(operator="equal", formula=['"完了"'], fill=green_fill),
    )
    ws.conditional_formatting.add(
        "G2:G1048576",
        CellIsRule(operator="equal", formula=['"遅延"'], fill=red_fill),
    )

    # ── オートフィルター ──
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"

    # ── 行の高さ ──
    ws.row_dimensions[1].height = 25
    for r in range(2, len(sample_data) + 2):
        ws.row_dimensions[r].height = 22

    # ── 保存 ──
    os.makedirs(config.DATA_DIR, exist_ok=True)
    wb.save(config.EXCEL_PATH)
    print(f"テンプレート作成完了: {config.EXCEL_PATH}")


if __name__ == "__main__":
    create_template()
